# -*- coding: utf-8 -*-
"""
docs/ROADMAP.md 기준 WBS(작업분해구조) 일정표 생성 스크립트
- 일반적 프로젝트 수행 방법론 기준의 계획형 WBS
- 3계층: 단계그룹(Level 1) → Phase(Level 2) → 작업패키지/TASK(Level 3)
- 영업일(평일) 누적으로 시작/종료일 자동 산출
- 산출: docs/wbs.xlsx (시트 3종)
출력: openpyxl
"""
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# 0. 일정 산정 기본 가정
# ──────────────────────────────────────────────────────────────
PROJECT_START = date(2026, 3, 2)  # 월요일, 프로젝트 착수 가정일

def add_workdays(start: date, days: int) -> date:
    """start 포함, days 영업일을 소요했을 때의 종료일(주말 제외)"""
    d = start
    cnt = 1
    while cnt < days:
        d += timedelta(days=1)
        if d.weekday() < 5:
            cnt += 1
    return d

def next_workday(d: date) -> date:
    d += timedelta(days=1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

# ──────────────────────────────────────────────────────────────
# 1. WBS 데이터 (ROADMAP.md 기준)
#    task = (TASK ID, 작업명, 상태, 담당역할, 산출물, 공수(영업일))
# ──────────────────────────────────────────────────────────────
DONE = "완료"
PLAN = "기획완료"
HOLD = "보류"
TODO = "예정"

groups = [
    {
        "name": "A. 기반 플랫폼 구축",
        "desc": "인증·결제·관리자·게시판·데이터표준·다국어 등 서비스 토대",
        "phases": [
            {"no": "Phase 0", "title": "스타터킷 현행화", "status": DONE, "tasks": [
                ("TASK-001", "Next.js 16 + Tailwind v4 + shadcn base-nova 환경 셋업", DONE, "DevOps", "재사용 스타터킷·Vercel 배포 파이프라인", 4),
            ]},
            {"no": "Phase 1", "title": "Pi Network 인증 + 결제", "status": DONE, "tasks": [
                ("TASK-002", "Pi 계정 인증 (HMAC 세션 쿠키)", DONE, "인증", "pi-auth-provider·/api/auth/pi", 4),
                ("TASK-003", "Pi Coin 결제 U2A 3단계 흐름", DONE, "인증", "approve/complete API·pi-pay-button", 4),
            ]},
            {"no": "Phase 2", "title": "Google 로그인 + 계정 연동", "status": DONE, "tasks": [
                ("TASK-004", "Supabase DB 설계 + NextAuth.js 연동", DONE, "BE", "sys_user·auth.ts·supabase-admin", 3),
                ("TASK-005", "Google 로그인 UI", DONE, "FE", "google-login-button·user-card", 2),
                ("TASK-006", "Pi Browser 감지 안정화", DONE, "인증", "authenticate 기반 분기 로직", 2),
                ("TASK-007", "Pi + Google 계정 연동 (6자리 OTP)", DONE, "인증", "link-start/complete/status API·link 페이지", 4),
                ("TASK-008", "브라우저별 UI 정리", DONE, "FE", "환경별 조건부 렌더링", 1),
            ]},
            {"no": "Phase 3", "title": "관리자 기능", "status": DONE, "tasks": [
                ("TASK-009", "관리자 기반 구조 (역할 접근제어)", DONE, "BE", "(admin) 레이아웃·사이드바·users 관리", 3),
                ("TASK-010", "결제 내역 관리", DONE, "BE", "admin/payments·π 합계", 2),
                ("TASK-011", "계정 연동 현황", DONE, "BE", "admin/links 통계", 2),
            ]},
            {"no": "Phase 4", "title": "통합 게시판", "status": DONE, "tasks": [
                ("TASK-020", "DB 스키마 + 기반 구조 (brd_*)", DONE, "DA", "brd_ctgr/post/cmnt/attch 설계", 2),
                ("TASK-021", "게시글 CRUD API", DONE, "BE", "/api/board 목록·상세·작성·수정·삭제", 3),
                ("TASK-022", "댓글 + QNA 채택 API", DONE, "BE", "comments·accept 엔드포인트", 2),
                ("TASK-023", "첨부파일 API (Supabase Storage)", DONE, "BE", "다중 업로드·롤백 처리", 2),
                ("TASK-024", "게시판 UI", DONE, "FE", "목록·작성·상세·수정 페이지", 4),
                ("TASK-025", "관리자 게시판 관리", DONE, "FE", "admin/board 핀·삭제 UI", 2),
            ]},
            {"no": "Phase 5", "title": "데이터 표준 시스템", "status": DONE, "tasks": [
                ("DA-QA", "DA 품질 점검 표준화 (스키마 정비)", DONE, "DA", "Migration 003~008·표준규칙 문서", 5),
                ("TASK-030", "표준단어 관리 (std_dic)", DONE, "DA", "words API·UI", 2),
                ("TASK-031", "표준도메인 관리 (std_dom)", DONE, "DA", "domains API·UI", 2),
                ("TASK-032", "표준용어 관리 (std_term)", DONE, "DA", "terms API·물리명 자동생성", 3),
                ("TASK-033", "DDL Export", DONE, "DA", "CREATE TABLE 자동 생성·다운로드", 2),
                ("TASK-034", "Audit Trail 변경 이력", DONE, "DA", "std_audit_log·트리거 함수", 2),
                ("TASK-035", "승인 워크플로우", DONE, "DA", "approval_queue·승인/반려 API", 3),
            ]},
            {"no": "Phase 6", "title": "다국어 처리 (next-intl v4)", "status": DONE, "tasks": [
                ("TASK-040", "next-intl v4 설치 + [locale] 라우팅", DONE, "i18n", "middleware·routing.ts", 3),
                ("TASK-041", "번역 파일 + UI 적용 (409키)", DONE, "i18n", "ko.json·3단계 fallback", 4),
                ("TASK-042", "국가/언어 DB + 동기화", DONE, "BE", "i18n_locale/message·sync API", 3),
                ("TASK-043", "AI 자동 번역 (Gemini Flash)", DONE, "BE", "배치 번역·18개 언어", 3),
                ("TASK-044", "다국어 안정성 강화 + 보안 패치", DONE, "BE", "단일소스·203 locale 선점·인젝션 방지", 3),
            ]},
        ],
    },
    {
        "name": "B. 핵심 서비스 — PiCafé",
        "desc": "테마 기반 카페·수익화·생태계 확장",
        "phases": [
            {"no": "Phase 7", "title": "PiCafé MVP", "status": DONE, "tasks": [
                ("TASK-050", "DB 마이그레이션 (msg_* 13개 테이블)", DONE, "DA", "sql/012_msg_tables.sql", 3),
                ("TASK-051", "테마 마스터 데이터 세팅", DONE, "BE", "테마·구독플랜·스티커 시드", 2),
                ("TASK-052", "1:1 카페 API + Supabase Realtime", DONE, "BE", "메시지 송수신·실시간 구독", 4),
                ("TASK-053", "그룹 카페 생성 (테마 선택 + Pi 결제)", DONE, "FE", "생성 UX·결제 연동", 4),
                ("TASK-055", "Pi Browser 쿠키 비의존 인증 (X-Pi-Token)", DONE, "인증", "헤더 폴백·piFetch", 3),
                ("TASK-054", "구독 시스템 (플랜 + Pi 결제 + PiRC2)", DONE, "BE", "구독 결제·Soroban 연동", 5),
            ]},
            {"no": "Phase 8", "title": "PiCafé 수익화 기능", "status": DONE, "tasks": [
                ("TASK-060", "Pi Tip (인라인 결제 + TIP_NOTI)", DONE, "BE", "팁 결제·알림 메시지", 3),
                ("TASK-061", "스티커 마켓 (테마별 팩 + 업셀)", DONE, "FE", "스티커 구매·업셀 UI", 3),
                ("TASK-062", "인라인 구매 트리거 8종", DONE, "FE", "구매 트리거 컴포넌트", 3),
                ("TASK-063", "이벤트 카페 (유료 입장 + 수익 분배)", DONE, "BE", "입장료·방장 정산", 3),
                ("TASK-064", "AI 카페 비서 (@ai 멘션)", DONE, "BE", "테마별 프롬프트·응답", 3),
                ("TASK-065", "파일·이미지·음성 메시지", DONE, "BE", "Storage 업로드·미디어 메시지", 3),
            ]},
            {"no": "Phase 9", "title": "PiCafé 생태계 확장", "status": DONE, "tasks": [
                ("TASK-070", "카페 마켓플레이스 (공개방 디렉토리)", DONE, "FE", "테마별 디렉토리", 3),
                ("TASK-071", "Pi Bet 투표", DONE, "BE", "투표·정산 로직", 3),
                ("TASK-072", "카페 봇·Webhook 연동 (Business)", DONE, "BE", "Webhook 엔드포인트", 3),
                ("TASK-073", "분석 대시보드 (Business)", DONE, "FE", "카페 분석 화면", 3),
                ("TASK-074", "커스텀 스티커 제작 (Business)", DONE, "FE", "스티커 업로드·제작", 3),
                ("P9-후속", "Pi Bet UI 개선 (아코디언)", DONE, "FE", "Bet UI 개선", 1),
            ]},
        ],
    },
    {
        "name": "C. 사용자 경험 · 운영",
        "desc": "마이페이지·통계 대시보드",
        "phases": [
            {"no": "Phase 10", "title": "사용자 프로필 관리 (마이페이지)", "status": DONE, "tasks": [
                ("TASK-056", "DB 마이그레이션 — sys_user 프로필 컬럼", DONE, "DA", "프로필 컬럼 추가", 1),
                ("TASK-057", "API — GET /api/profile", DONE, "BE", "프로필 조회", 1),
                ("TASK-058", "API — PATCH /api/profile", DONE, "BE", "프로필 수정", 1),
                ("TASK-059", "API — GET /api/profile/payments", DONE, "BE", "결제 내역 조회", 1),
                ("TASK-060p", "컴포넌트 — ProfileTabs + ClientProfileGate", DONE, "FE", "탭·클라이언트 게이트", 2),
                ("TASK-061p", "컴포넌트 — ProfileForm·결제·구독현황", DONE, "FE", "프로필 폼·결제·구독 표시", 3),
                ("TASK-062p", "번역 + 3단계 검증", DONE, "i18n", "다국어·검증", 1),
            ]},
            {"no": "Phase 11", "title": "어드민 통계 대시보드", "status": DONE, "tasks": [
                ("TASK-080", "활동 로그 마이그레이션", DONE, "DA", "sql/015 활동 로그", 1),
                ("TASK-081", "활동 계측 (원천 적재)", DONE, "BE", "이벤트 계측 코드", 2),
                ("TASK-082", "중간집계 테이블 + 집계 RPC", DONE, "DA", "sql/016 집계", 2),
                ("TASK-083", "집계 배치 + 백필", DONE, "BE", "배치·과거 데이터 백필", 2),
                ("TASK-084", "통계 API", DONE, "BE", "DAU/WAU/MAU·매출 API", 2),
                ("TASK-085", "차트 라이브러리 + 컴포넌트 3종", DONE, "FE", "차트 컴포넌트", 2),
                ("TASK-086", "대시보드 페이지 + 메뉴", DONE, "FE", "통계 대시보드 화면", 2),
                ("TASK-087", "검증", DONE, "QA", "통계 정합성 검증", 1),
                ("P11-후속1", "후속 고도화 1차", DONE, "FE", "대시보드 개선", 1),
                ("P11-후속2", "후속 고도화 2차 (트리맵·KST 교정)", DONE, "FE", "coin360 트리맵·집계 교정", 2),
            ]},
        ],
    },
    {
        "name": "D. 글로벌 · 커머스 확장",
        "desc": "실시간 동시통역·P2P 직거래 마켓",
        "phases": [
            {"no": "Phase 12", "title": "PiTranslate™ 글로벌 동시통역", "status": DONE, "tasks": [
                ("TASK-090", "DB 마이그레이션 + 환경변수", DONE, "DA", "번역 캐시 테이블", 1),
                ("TASK-091", "Gemini Flash 번역 라이브러리", DONE, "BE", "번역 lib", 2),
                ("TASK-092", "동시성 dedup 처리", DONE, "BE", "중복 요청 제거", 2),
                ("TASK-093", "번역 API 라우트", DONE, "BE", "translate API", 2),
                ("TASK-094", "메시지 전송 시 번역 큐 연동", DONE, "BE", "큐 연동", 2),
                ("TASK-095", "클라이언트 broadcast 구독 확장", DONE, "FE", "실시간 번역 수신", 2),
                ("TASK-096", "사용자 표시 언어 설정 UI", DONE, "FE", "언어 설정", 1),
                ("TASK-097", "원문 보기 토글 UI", DONE, "FE", "원문/번역 토글", 1),
                ("TASK-098", "어드민 번역 통계", DONE, "FE", "번역 통계 화면", 1),
                ("TASK-099", "번역 품질 피드백", DONE, "BE", "피드백 수집", 1),
            ]},
            {"no": "Phase 13", "title": "PiShop(MPS) P2P 직거래", "status": DONE, "tasks": [
                ("TASK-100", "DB 마이그레이션 sql/029_mps.sql", DONE, "DA", "MPS 테이블", 2),
                ("TASK-101", "lib 헬퍼 3종", DONE, "BE", "MPS 공통 헬퍼", 1),
                ("TASK-102", "상품 API (등록·조회·이미지)", DONE, "BE", "상품 CRUD·이미지 업로드", 3),
                ("TASK-103", "재고 관리 (RPC 단일 트랜잭션)", DONE, "BE", "재고 차감·센티널", 2),
                ("TASK-104", "주문 + 에스크로 API (양방향 확인)", DONE, "BE", "주문·에스크로·취소", 4),
                ("TASK-105", "상품 목록·상세 UI", DONE, "FE", "SCR-01/02", 3),
                ("TASK-106", "내 상품 관리 UI", DONE, "FE", "SCR-03/04·등록 폼", 3),
                ("TASK-107", "주문 관리 UI", DONE, "FE", "SCR-05/06", 2),
                ("TASK-108", "카테고리 시스템 (FR-03)", DONE, "BE", "카테고리 관리", 2),
                ("TASK-109", "매장 관리 (FR-06)", DONE, "FE", "매장 관리·SCR-08", 2),
                ("TASK-110", "양방향 주문 취소 (FR-10)", DONE, "BE", "취소 워크플로우", 2),
                ("TASK-111", "거래 내역 (FR-12·SCR-07)", DONE, "FE", "거래 내역 화면", 2),
                ("P13-후속", "후속 개선 (A2U 자동환불·이미지·취소 UI)", DONE, "BE", "환불·UI 개선", 2),
                ("TASK-113", "Google Maps 연동", TODO, "FE", "Maps API 키 필요·예정", 3),
                ("TASK-112", "PiRC3 실 에스크로 마이그레이션", HOLD, "BE", "선결조건 미충족·보류", 0),
            ]},
        ],
    },
    {
        "name": "E. 실시간 · 위치 서비스",
        "desc": "WebRTC 음성채널·위치기반서비스(LBS)",
        "phases": [
            {"no": "Phase 14", "title": "PiVoice™ WebRTC N:N 음성채널", "status": DONE, "tasks": [
                ("TASK-120", "데이터 모델 sql/032_voice_channel", DONE, "DA", "음성채널 테이블", 1),
                ("TASK-121", "TURN 자격증명 발급 API", DONE, "BE", "TTL 자격증명", 2),
                ("TASK-122", "음성채널 API 5종", DONE, "BE", "채널 입퇴장·관리 API", 3),
                ("TASK-123", "WebRTC 훅 + UI", DONE, "FE", "Full Mesh 연결·UI", 5),
                ("TASK-124", "v3.0 권한 시스템 (보장슬롯·승인)", DONE, "BE", "방장 슬롯·발언 승인 R1~R7", 3),
                ("TASK-125", "S0 진단 메시지 (입장 실패 사유)", DONE, "FE", "진단 화면", 1),
                ("PV-잔여", "S0 실기기 검증 · TURN 운영 설정", TODO, "QA", "Pi Browser 마이크 검증·TURN env", 3),
            ]},
            {"no": "Phase 15", "title": "LBS 위치기반서비스", "status": DONE, "tasks": [
                ("TASK-130", "DB 마이그레이션 sql/033_lbs.sql", DONE, "DA", "위치 테이블", 1),
                ("TASK-131", "환경변수 + Google Maps API 설정", DONE, "DevOps", "Maps API 설정", 1),
                ("TASK-132", "동의 API (consent)", DONE, "BE", "위치 동의 GET/POST/DELETE", 1),
                ("TASK-133", "위치 저장 API (save)", DONE, "BE", "위치 저장", 1),
                ("TASK-134", "Google Maps 서버 프록시 API", DONE, "BE", "Maps 프록시", 2),
                ("TASK-135", "주변 탐색 API (nearby)", DONE, "BE", "거리 기반 탐색", 2),
                ("TASK-136", "MPS 상품 목록 거리 표시", DONE, "BE", "거리 계산·표시", 1),
                ("TASK-137", "클라이언트 동의 플로우 UI", DONE, "FE", "동의 UI", 2),
                ("TASK-138", "상품 카드 거리 UI + 반경 필터", DONE, "FE", "거리·반경 필터", 2),
                ("TASK-139", "touchLastLogin 연동 + GPS 검증", DONE, "BE", "로그인 위치·GPS 검증", 1),
                ("TASK-140", "상품 개별 위치 등록 (판매자 GPS)", DONE, "FE", "등록 시 GPS 저장·거리 기준 전환", 2),
            ]},
        ],
    },
    {
        "name": "F. 이벤트 · 마케팅",
        "desc": "Pi 요원 육성 이벤트 미션 시스템",
        "phases": [
            {"no": "Phase 16", "title": "이벤트 미션 시스템 (구현 대기)", "status": PLAN, "tasks": [
                ("TASK-150", "이벤트 DB 스키마", PLAN, "DA", "이벤트·미션·화이트리스트 테이블", 2),
                ("TASK-151", "미션 완료 자동 감지 훅", PLAN, "BE", "미션 트래킹 훅", 3),
                ("TASK-152", "Footer Event 탭 + 이벤트 페이지", PLAN, "FE", "이벤트 페이지·10미션 UI", 4),
                ("TASK-153", "랭킹 + 선착순 선물 관리", PLAN, "BE", "랭킹·선착순 로직", 3),
                ("TASK-154", "관리자 제외 관리", PLAN, "FE", "제외 대상 관리 UI", 2),
            ]},
        ],
    },
    {
        "name": "G. 횡단 개선 (성능·안정화)",
        "desc": "전체 기능 대상 성능 튜닝·Pi Browser 안정화",
        "phases": [
            {"no": "횡단 1", "title": "성능 튜닝", "status": DONE, "tasks": [
                ("X1", "무한 스크롤·지연 로딩 + SWR 캐싱·병렬 호출·리브랜딩", DONE, "FE", "성능 튜닝·스티커 노출 개선", 3),
            ]},
            {"no": "횡단 3", "title": "Pi Browser 안정화·콤보 성능", "status": DONE, "tasks": [
                ("X3", "admin 다국어 전환 무반응 수정 + 헤더 콤보 3계층 캐시", DONE, "FE", "_pit 티켓 선발급·캐시", 2),
            ]},
            {"no": "횡단 4", "title": "안정화 4차·대시보드 고도화", "status": DONE, "tasks": [
                ("X4", "헤더 로고·다국어 기억·브랜드 통일·open redirect 방어·대시보드 고도화", DONE, "FE", "안정화·UI 통일·보안", 3),
            ]},
        ],
    },
]

# ──────────────────────────────────────────────────────────────
# 2. 일정 계산 (영업일 순차 누적) + 평탄화
# ──────────────────────────────────────────────────────────────
PROGRESS = {DONE: 100, PLAN: 10, HOLD: 0, TODO: 0}

rows = []  # (level, wbs, name, status, progress, role, deliverable, days, start, end, pred, note)
cursor = PROJECT_START
prev_task_wbs = ""

for gi, g in enumerate(groups, start=1):
    g_wbs = str(gi)
    g_start = None
    g_end = None
    g_days = 0
    g_rows_start_idx = len(rows)
    rows.append(["L1", g_wbs, g["name"], "", "", "", g["desc"], None, None, None, "", ""])
    g_row_idx = len(rows) - 1

    for pi_, p in enumerate(g["phases"], start=1):
        p_wbs = f"{g_wbs}.{pi_}"
        p_start = None
        p_end = None
        p_days = 0
        rows.append(["L2", p_wbs, f"{p['no']} — {p['title']}", p["status"], "", "", "", None, None, None, "", ""])
        p_row_idx = len(rows) - 1

        for ti, t in enumerate(g["phases"][pi_-1]["tasks"], start=1):
            tid, tname, tstatus, trole, tdeliv, tdays = t
            t_wbs = f"{p_wbs}.{ti}"
            if tdays > 0:
                t_start = cursor
                t_end = add_workdays(t_start, tdays)
                cursor = next_workday(t_end)
            else:
                # 보류 작업: 일정 미배정
                t_start = None
                t_end = None
            note = f"{tid}"
            pred = prev_task_wbs if tdays > 0 else ""
            rows.append([
                "L3", t_wbs, f"{tid}: {tname}", tstatus, PROGRESS[tstatus],
                trole, tdeliv, (tdays if tdays > 0 else None), t_start, t_end, pred, ""
            ])
            if tdays > 0:
                prev_task_wbs = t_wbs
                p_days += tdays
                if p_start is None or t_start < p_start:
                    p_start = t_start
                if p_end is None or t_end > p_end:
                    p_end = t_end

        # Phase roll-up
        rows[p_row_idx][7] = p_days if p_days else None
        rows[p_row_idx][8] = p_start
        rows[p_row_idx][9] = p_end
        # Phase 진행률 = 하위 평균
        sub = [r for r in rows[p_row_idx+1:] if r[0] == "L3"]
        sub = [r for r in rows if r[0] == "L3" and r[1].startswith(p_wbs + ".")]
        if sub:
            rows[p_row_idx][4] = round(sum(s[4] for s in sub) / len(sub))
        if p_days:
            g_days += p_days
            if g_start is None or (p_start and p_start < g_start):
                g_start = p_start
            if g_end is None or (p_end and p_end > g_end):
                g_end = p_end

    # Group roll-up
    rows[g_row_idx][7] = g_days if g_days else None
    rows[g_row_idx][8] = g_start
    rows[g_row_idx][9] = g_end
    gsub = [r for r in rows if r[0] == "L3" and r[1].startswith(g_wbs + ".")]
    if gsub:
        rows[g_row_idx][4] = round(sum(s[4] for s in gsub) / len(gsub))

PROJECT_END = cursor  # 마지막 작업 다음 영업일 (직전이 실제 종료)

# ──────────────────────────────────────────────────────────────
# 3. 엑셀 작성
# ──────────────────────────────────────────────────────────────
wb = Workbook()

# 색상/스타일 정의
C_HEAD = PatternFill("solid", fgColor="1F4E78")
C_L1   = PatternFill("solid", fgColor="2E75B6")
C_L2   = PatternFill("solid", fgColor="BDD7EE")
C_DONE = PatternFill("solid", fgColor="C6EFCE")
C_PLAN = PatternFill("solid", fgColor="FFEB9C")
C_HOLD = PatternFill("solid", fgColor="FFC7CE")
C_TODO = PatternFill("solid", fgColor="FCE4D6")
F_HEAD = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
F_L1   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
F_L2   = Font(name="맑은 고딕", bold=True, color="1F4E78", size=10)
F_BASE = Font(name="맑은 고딕", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def status_fill(s):
    return {DONE: C_DONE, PLAN: C_PLAN, HOLD: C_HOLD, TODO: C_TODO}.get(s)

# ── 시트 1: 프로젝트 개요 ────────────────────────────────────
ws0 = wb.active
ws0.title = "프로젝트 개요"
overview = [
    ("Pi Network 기반 풀스택 앱 플랫폼 — WBS 일정표", "title"),
    ("", ""),
    ("문서 정보", "h"),
    ("기준 문서", "docs/ROADMAP.md (기준일 2026-06-14)"),
    ("작성 방법론", "WBS(작업분해구조) 3계층 — 단계그룹 → Phase → 작업패키지(TASK)"),
    ("일정 산정 방식", "영업일(평일) 순차 누적 / 주말 제외 / 1인~소수 인력 순차 진행 가정"),
    ("프로젝트 착수 가정일", PROJECT_START.strftime("%Y-%m-%d (월)")),
    ("프로젝트 종료 예정", add_workdays(PROJECT_START, 1) and "", ),  # placeholder
    ("", ""),
    ("진행 현황 요약", "h"),
]
# 통계 계산
all_tasks = [r for r in rows if r[0] == "L3"]
cnt_done = sum(1 for r in all_tasks if r[3] == DONE)
cnt_plan = sum(1 for r in all_tasks if r[3] == PLAN)
cnt_hold = sum(1 for r in all_tasks if r[3] == HOLD)
cnt_todo = sum(1 for r in all_tasks if r[3] == TODO)
total = len(all_tasks)
# 실제 종료일 = 일정 배정된 작업 중 최대 end
ends = [r[9] for r in all_tasks if r[9]]
real_end = max(ends)
overview[7] = ("프로젝트 종료 예정", real_end.strftime("%Y-%m-%d"))

stat_rows = [
    ("전체 작업패키지(TASK)", f"{total} 건"),
    ("완료(✅)", f"{cnt_done} 건"),
    ("기획완료·구현대기(📋)", f"{cnt_plan} 건"),
    ("예정(🔜)", f"{cnt_todo} 건"),
    ("보류(🔒)", f"{cnt_hold} 건"),
    ("전체 진행률", f"{round(sum(r[4] for r in all_tasks)/total)}%"),
    ("총 단계그룹", f"{len(groups)} 개"),
    ("총 Phase", f"{sum(len(g['phases']) for g in groups)} 개"),
    ("배정 총 공수", f"{sum(r[7] for r in all_tasks if r[7])} 영업일"),
]
legend_rows = [
    ("", ""),
    ("범례 (상태)", "h"),
    ("완료", "구현·검증 완료된 작업"),
    ("기획완료", "PRD 등 기획 완료, 구현 대기"),
    ("예정", "후속 예정 작업"),
    ("보류", "선결조건 미충족으로 보류"),
    ("", ""),
    ("담당역할 약어", "h"),
    ("기획 / PM", "제품 기획·일정 관리"),
    ("BE", "백엔드 (API·DB 로직)"),
    ("FE", "프론트엔드 (UI·클라이언트)"),
    ("DA", "데이터 아키텍트 (표준·스키마)"),
    ("인증", "Pi/Auth 인증 전문"),
    ("i18n", "다국어"),
    ("DevOps", "인프라·배포"),
    ("QA", "테스트·검증"),
]

r_idx = 1
for label, val in overview + stat_rows + legend_rows:
    cell_a = ws0.cell(row=r_idx, column=1, value=label)
    cell_b = ws0.cell(row=r_idx, column=2, value=val if val != "h" and val != "title" else "")
    if val == "title":
        cell_a.font = Font(name="맑은 고딕", bold=True, size=16, color="1F4E78")
    elif val == "h":
        cell_a.font = Font(name="맑은 고딕", bold=True, size=12, color="FFFFFF")
        cell_a.fill = C_L1
        ws0.cell(row=r_idx, column=2).fill = C_L1
    else:
        cell_a.font = Font(name="맑은 고딕", bold=True, size=10)
        cell_b.font = F_BASE
        # 상태 범례 색칠
        if label in (DONE, PLAN, TODO, HOLD):
            cell_a.fill = status_fill(label)
    r_idx += 1
ws0.column_dimensions["A"].width = 26
ws0.column_dimensions["B"].width = 70

# ── 시트 2: WBS 일정표 ──────────────────────────────────────
ws = wb.create_sheet("WBS 일정표")
headers = ["WBS", "구분", "작업명", "상태", "진행률", "담당", "산출물", "공수(일)", "시작일", "종료일", "선행", "비고"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = C_HEAD
    cell.font = F_HEAD
    cell.alignment = CENTER
    cell.border = BORDER

lvl_label = {"L1": "단계그룹", "L2": "Phase", "L3": "TASK"}
for i, r in enumerate(rows, start=2):
    level, wbs, name, status, prog, role, deliv, days, start, end, pred, note = r
    vals = [
        wbs, lvl_label[level], name, status,
        (f"{prog}%" if prog != "" else ""),
        role, deliv, days,
        (start.strftime("%Y-%m-%d") if start else ""),
        (end.strftime("%Y-%m-%d") if end else ""),
        pred, note,
    ]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.border = BORDER
        cell.font = F_BASE
        if c in (1, 2, 4, 5, 6, 8, 9, 10, 11):
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT
    # 레벨별 행 스타일
    if level == "L1":
        for c in range(1, 13):
            ws.cell(row=i, column=c).fill = C_L1
            ws.cell(row=i, column=c).font = F_L1
    elif level == "L2":
        for c in range(1, 13):
            ws.cell(row=i, column=c).fill = C_L2
            ws.cell(row=i, column=c).font = F_L2
    else:
        # 상태 색칠 (상태 칸만)
        sf = status_fill(status)
        if sf:
            ws.cell(row=i, column=4).fill = sf

# 열 너비
widths = [9, 9, 46, 9, 8, 7, 34, 8, 12, 12, 8, 8]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{len(rows)+1}"

# ── 시트 3: Phase 요약 (마일스톤 + 텍스트 간트) ──────────────
ws2 = wb.create_sheet("Phase 요약")
phase_rows = [r for r in rows if r[0] == "L2"]
# 간트 타임라인 범위 (월 단위)
gantt_start = min(r[8] for r in phase_rows if r[8])
gantt_end = max(r[9] for r in phase_rows if r[9])
# 월 목록 생성
months = []
y, m = gantt_start.year, gantt_start.month
while (y, m) <= (gantt_end.year, gantt_end.month):
    months.append((y, m))
    m += 1
    if m > 12:
        m = 1; y += 1

head2 = ["Phase", "상태", "진행률", "공수(일)", "시작일", "종료일"] + [f"{yy%100:02d}/{mm:02d}" for yy, mm in months]
for c, h in enumerate(head2, start=1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.fill = C_HEAD
    cell.font = F_HEAD
    cell.alignment = CENTER
    cell.border = BORDER

C_BAR = PatternFill("solid", fgColor="2E75B6")
C_BAR_PLAN = PatternFill("solid", fgColor="FFC000")
for i, r in enumerate(phase_rows, start=2):
    _, wbs, name, status, prog, _, _, days, start, end, _, _ = r
    base = [name, status, f"{prog}%", days,
            start.strftime("%Y-%m-%d") if start else "",
            end.strftime("%Y-%m-%d") if end else ""]
    for c, v in enumerate(base, start=1):
        cell = ws2.cell(row=i, column=c, value=v)
        cell.border = BORDER
        cell.font = F_BASE
        cell.alignment = CENTER if c != 1 else LEFT
        if c == 2:
            sf = status_fill(status)
            if sf: cell.fill = sf
    # 간트 막대
    if start and end:
        for mi, (yy, mm) in enumerate(months):
            col = 7 + mi
            month_first = date(yy, mm, 1)
            month_last = date(yy + (1 if mm == 12 else 0), 1 if mm == 12 else mm + 1, 1) - timedelta(days=1)
            if start <= month_last and end >= month_first:
                cell = ws2.cell(row=i, column=col, value="")
                cell.fill = C_BAR_PLAN if status in (PLAN, TODO) else C_BAR
                cell.border = BORDER

ws2.column_dimensions["A"].width = 38
for c in range(2, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 11
for c in range(7, 7 + len(months)):
    ws2.column_dimensions[get_column_letter(c)].width = 6
ws2.freeze_panes = "B2"

# ──────────────────────────────────────────────────────────────
import os
out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs", "wbs.xlsx")
wb.save(out)
print("저장 완료:", out)
print(f"작업 {total}건 (완료 {cnt_done}/기획 {cnt_plan}/예정 {cnt_todo}/보류 {cnt_hold})")
print(f"일정: {PROJECT_START} ~ {real_end} (배정 공수 {sum(r[7] for r in all_tasks if r[7])} 영업일)")
print(f"Phase {len(phase_rows)}개, 간트 {len(months)}개월")
